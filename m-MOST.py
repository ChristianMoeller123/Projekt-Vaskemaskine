import math
import pickle  # Read/write objects
import re  # Library for handling strings
import os
import openpyxl
import sys

with open('objects.pickle', 'rb') as f:
    x = pickle.load(f)
    AllParents = x['Parents']
    AllChildren = x['Children']
    AllActions = x['Actions']
    AllDisassemblies = x['Disassemblies']
    AllPACUnits = x['PACUnits']

#  Distances
try:
    Atemp = int(os.environ['A'])
    if Atemp in [1,2]:
        A = 30
    elif Atemp in [3,4]:
        A = 60
    elif Atemp in [5,6,7]:
        A = 100
    elif Atemp in [8,9,10]:
        A = 160
    else:
        print("Consider putting you tables closer to the setup!")
except:
    A = 30
Ab = A  # Distance from disassembly to desk 2. INDEX!
Aa = Ab
Ac = Ab
Ad = Ab
Ai = math.sqrt(Ab^2 + Ad^2)
Ae = 2*Ai
Ah = Ai
Af =  100  #  GET INFORMATION FROM USER!!! From desk1 to storage
Ag =  100  #  GET INFORMATION FROM USER!!! from strage to cross
#  Idea: Regression from table --> function between meters and index --> input meters get index
Wa = 5000 #  Rough asumption about the time it takes to input in excel/input gui




#  Function to return star.
def GaFun(ToolInput):
    # Takes care if there are more tools represented
    #  Split each tool in the string
    Tools = re.split(r'&', ToolInput)
    #  Strip for white spaces
    Tools = [Tool.strip() for Tool in Tools]
    #  Remove empty strings from the list
    Tools = list(filter(None, Tools))

    Ga = 0
    for Tool in Tools:
        Tool = Tool.lower()
        # Small tools
        if Tool in ['screwdriver', 'hand', 'hands', 'plectrum', 'wrench', 'hammer', 'cross-head screwdriver', 'socket wrench', 'tri-wing screwdriver', 'flat head screwdriver', 'flathead screwdriver', 'wirecutter']:
            Ga += 10

        # Big tools
        elif Tool in ['electric screwdriver', 'drill', 'electric screwdriver']:
            Ga += 30
        else:
            print('Given tool: '+Tool+' is not in our definitions (GA)')
            return
    return Ga

def StarFun(ToolInput, ActionNameInput):
    # Takes care if there are more tools represented
    #  Split each tool in the string
    Tools = re.split(r'&'r','r'.', ToolInput)
    #  Strip for white spaces
    Tools = [Tool.strip() for Tool in Tools]
    #  Remove empty strings from the list
    Tools = list(filter(None, Tools))

    #  Same for action name
    ActionNames = re.split(r',|&', ActionNameInput)
    ActionNames = [actionname.strip() for actionname in ActionNames]
    ActionNames = list(filter(None, ActionNames))

    Star = 0
    for Tool in Tools:
        Tool = Tool.lower()
        if Tool in ['screwdriver', 'tri-wing screwdriver', 'flat head screwdriver', 'cross-head screwdriver', 'wirecutter', 'flathead screwdriver']:  # Assume 9 turns with screwdriver
            Star += 160
        elif Tool in ['hand', 'hands', 'plectrum']:  #  Hand depends on action name
            #  Loop over all action names. This could definetly be improved regarding assumptions!
            for ActionName in ActionNames:
                ActionName = ActionName.lower()
                if ActionName in ['remove', 'disconnect']:
                    Star += 60  # two hands one turn
                elif ActionName == 'separate':
                    Star += 160  #  2-hands 3 turns
        elif Tool in ['electric screwdriver', 'angle grinder', 'electric screwdriver']:  #  Power tools
            Star += 30  #  Assume screw diam. of 6mm
        elif Tool in ['socket wrench', 'wrench']: # Assume 5 turns with wrench
            Star += 100
        elif Tool == 'hammer':
            Star += 60  #  Assume 3 strikes or 6 taps with hammer
        else:
            print('Given tool: ' + Tool + ' is not in our definitions (Star)')
            return
    return Star



#  INDEX TO S function. Takes index, gives seconds
def indexSeqToS(Sequence):
    time = sum(Sequence)*0.036  #TMU*0.036 = 1 second
    time = float(time)
    return round(time, 2)




# Base MOST for Action tools (and times)
def DEIAction(action, AllPACUnits):
    tool = action.Tool
    times = action.Times
    ActionName = action.Desc
    #  Variables for MOST

    action.DEISteps = []
    #  Steps 1-2
    Sequence = [Aa, 50, 10, 30, 50, 10]
    action.DEISteps.append(indexSeqToS(Sequence))

    #  Steps 3-5
    try:
        times = int(times)
        Ga = GaFun(tool)
        Star = StarFun(tool, ActionName)
        # Repeat the sequence the specified number of times
        Sequence = [Ab, Ga, Ab, 10, 30, Star, Ab, 10, 10]
        if times > 1:
            for i in range(1, times):
                Sequence.extend([10, 30, Star, 10, 10])
    except:
        # Parse the string and generate the corresponding sequence for each action
        steps = times.split('.')
        for i in range(0, len(steps)):
            split_steps = steps[i].split()
            tool_times = int(split_steps[0])
            tool_desc = split_steps[1].lower()
            if tool_desc == "remove":
                actionname = "hands"
            elif tool_desc == "cut":
                actionname = "wirecutter"
            elif tool_desc == "unscrew":
                actionname = "electric screwdriver"
            Star = StarFun(actionname, tool_desc) # i switched actioname and tool_desc here,  so actionname is actually the tool
            Ga = GaFun(actionname)
            times = int(tool_times)
            Sequence = [Ab, Ga, Ab, 10, 30, Star, Ab, 10, 10]
            if times > 1:
                for i in range(1, times):
                    Sequence.extend([10, 30, Star, 10, 10])

    action.DEISteps.append(indexSeqToS(Sequence))
    #  Step 6 More than 1 non-leaf, add somthing
    # Find children, and determine amount of non-leaf children:
    nLeafNodes = 0
    children = AllPACUnits[action.PACID-1].Children if isinstance(AllPACUnits[action.PACID-1].Children, list) else [AllPACUnits[action.PACID-1].Children]
    for j in range(len(children)):
        if children[j].isLeaf:
            nLeafNodes += 1
    Sequence = [Ab, 10, 10, Ac, 10, 10]
    if nLeafNodes > 1:
        Sequence.extend([Ac, 10, 30, Ad, 10, 10, Ae])
    action.DEISteps.append(indexSeqToS(Sequence))

    #  Step 7
    #  Remember very rough asumption here about input time!!!
    Sequence = [10, 10, Wa]
    action.DEISteps.append(indexSeqToS(Sequence))

    #  Step 8
    Sequence = [0]
    for i in range(0, nLeafNodes):
        Sequence.extend([10, 30, 10, 30, 160, 10, 30, 10, 10, 10, 30, 30, 10, 10])
        #Sequence.extend([10, 10, 10, 30, 10, 10, 10]) old sequence
    action.DEISteps.append(indexSeqToS(Sequence))

    #  Steps 9-10
    Sequence = [10, 10, Af, 10, 10, Ag]
    action.DEISteps.append(indexSeqToS(Sequence))

    '''
    If conditionen på steps 11-19  giver ikke mening
    #  Step 11-19 if there were more than 1 non leaf....
    #  continue this until only leaf children, but only do this for subassemblies --> WE NEED INPUT
    if nLeafNodes > 1:
        #  Step 11
        Sequence = [Ah, 50]
        action.DEISteps.append(indexSeqToS(Sequence))

        #  Steps 12-14
        Sequence = [Ai, Ga, Ai, 10, 30, Star, 10, 10]
        action.DEISteps.append(indexSeqToS(Sequence))

        #  Step 15
        Sequence = [Ai, 10, Ae, 10]
        action.DEISteps.append(indexSeqToS(Sequence))
        # SKER IKKE I GORENJE. TAG STILLING TIL HVIS DET SKER SENERE
        # DF For subassembly in this loop is different than prev. DF!

        # MANGLER STEPS 16-19!!!


    #  Start assumption: If DF then it's a big tool and only once
    #  From DF input, call steps 16-19
    '''

    DEItotal = sum(action.DEISteps)
    return round(DEItotal, 2)



for action in AllActions:
    # FIND DFs IN ACTION
    # Add them to the DEIAction call.
    #  Eller gør alt det inde i DEIAction???
    action.ActionDEI = DEIAction(action, AllPACUnits)
# TAGER IKKE HØJDE FOR N/A

# Loop over DFs and make that a different time parameter
# A final time parameter which is the sum of action and DF time.
for diss in AllDisassemblies:
    #  DFID definitions are different from XML and initial GUI!
    #  Initial GUI:
    origin = diss.DType.split("-")[2]

    #  XML:
    #origin = diss.DFID

    #  Only DFs on actions will have a DEI calculation:
    for char in origin:
        if char == 'A':
            index = origin.find('A')
            PACID = origin[:index]
            PACID = int(PACID)
            # DEI for DF in main disassembly (vaskemaskinen). Assume 1 action time, heavy tool
            Sequence = [Ab, 30, Ab, 10, 30, 30, Ab, 10, 10]
            #  Assume screw diam. of 6mm, and big electric tool for the star value.
            #  Assume big power tool for Ga.

            #  Assign the DEI to the action
            AllPACUnits[PACID-1].Action.DFDEI = indexSeqToS(Sequence)



#  Print MOST calculations to an excel sheet

workbook = openpyxl.Workbook()
worksheet = workbook.active

worksheet.append(["Action ID", "Steps 1-2", "Steps 3-5", "Step 6", "Step 7", "Step 8", "Steps 9-10", "Step 11", "Steps 12-14", "Step 15", "Steps 16-19", "DF"])



for i in range(len(AllActions)):
    worksheet.cell(row=i + 2, column=1, value=AllActions[i].ID)
    for j in range(len(AllActions[i].DEISteps)):
        worksheet.cell(row=i+2, column=j+2, value=AllActions[i].DEISteps[j])
        if AllActions[i].DFDEI:
            worksheet.cell(row=i + 2, column=12, value=AllActions[i].DFDEI)
workbook.save("GORENJE3_MOST_DATA.xlsx")



objects = {"Parents": AllParents, "Actions": AllActions, "Children": AllChildren, "Disassemblies": AllDisassemblies, "PACUnits": AllPACUnits}
# Deserializes the objects (find et link?)
with open('objects.pickle', 'wb') as f:
    pickle.dump(objects, f)
